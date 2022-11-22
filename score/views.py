from django.shortcuts import render,redirect
from .forms import CountForm
from .models import Run,Count,Score
from django.views.generic import TemplateView,CreateView,ListView,UpdateView,FormView
from django.urls import reverse_lazy
from django.http import HttpResponse
import openpyxl
import math
from operator import itemgetter
import base64
from django.core.files.base import ContentFile

# Create your views here.
#走行順にチーム名が並んでいるページ
#チーム名をクリックすると得点入力のページに移動する
class Index(TemplateView):
  template_name = 'score/index.html'
  model = Run
  def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        run = Run.objects.all().order_by("order")
        context['runs'] = run
        #context['form'] = CountForm
        return context

  def get(self, request, *args, **kwargs):
    return super().get(request, *args, **kwargs)

#得点入力のページに飛んできたときの処理
#Getの場合はフォームを表示する
#確認画面から戻ってくる場合はFormの値を保持したまま
class Create(FormView):
  template_name = "score/create.html"
  model = Count
  form_class = CountForm
  def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        runs = Run.objects.get(order = self.kwargs['pk'])
        context['runs'] = runs
        return context
      
  def get_form_kwargs(self):
        kwargs = super( Create, self).get_form_kwargs()
        if self.request.method=="GET":
            kwargs.update(self.request)
        return kwargs
  
  def get_initial(self):
    initial = super().get_initial()
    initial["run"] = self.kwargs['pk']
    return initial
  
  def form_valid(self, form):
    runs = Run.objects.get(order = self.kwargs['pk'])
    context ={
      'form':form,
      'runs':runs
    }
    return render(self.request, 'score/create.html',context)
  

#確認画面の処理
#得点入力画面で送信ボタンを押されたらこちらの処理に飛んでくる
#カウントから得点を計算して確認画面のテンプレートに送る
class Confirm(FormView):
  form_class = CountForm
  template_name = "score/confirm.html"
  #得点を計算するための自作関数
  def calc_score(self, form):
    m1_score = form.cleaned_data['m1_count'] * 100
    m2_score = form.cleaned_data["m2_count"] * 400
    m2h_score = form.cleaned_data["m2h_count"] * 200
    m3_score = form.cleaned_data["m3_count"] * 400
    m3h_score = form.cleaned_data["m3h_count"] * 200
    m4_score = form.cleaned_data["m4_count"] * 600
    m4h_score = form.cleaned_data["m4h_count"] * 300
    m6_score = form.cleaned_data["m6_count"] * 1000
    m6h_score = form.cleaned_data["m6h_count"] * 500
    bonus1_score = form.cleaned_data["bonus1_count"] * 200
    if(form.cleaned_data["m5_count"] == True):
      m5_score = 500
    else:
      m5_score = 0
      
    if(form.cleaned_data["m7_count"] == True):
      m7_score = 500
    else:
      m7_score = 0
      
    if(form.cleaned_data["bonus2_count"] == True):
      bonus2_score = 200
    else:
      bonus2_score = 0
    total = m1_score + m2_score + m2h_score +m3_score + m3h_score + m4_score + m4h_score + m6_score + m6h_score + m5_score +m7_score + bonus2_score + bonus1_score
    
    score = {
      'm1_score':m1_score,
      'm2_score':m2_score,
      'm2h_score':m2h_score,
      'm3_score':m3_score,
      'm3h_score':m3h_score,
      'm4_score':m4_score,
      'm4h_score':m4h_score,
      'm5_score':m5_score,
      'm6_score':m6_score,
      'm6h_score':m6h_score,
      'm7_score':m7_score,
      'bonus1_score':bonus1_score,
      'bonus2_score':bonus2_score,
      'total':total,
    }
    return score
  #フォームが送られたときに実行される
  #必要な変数をテンプレートへ送る     
  def form_valid(self, form):
    score = self.calc_score(form)
    pk = self.request.POST['run']
    run = Run.objects.get(order=pk)
    context ={
      'form':form,
      'pk':pk ,
      'run':run,
      'score':score,
      }
    return render(self.request, 'score/confirm.html',context)

  #フォームに何か問題があったときの処理
  #入力画面に戻る
  def form_invalid(self, form):
    msg ='コンテナの数が合いません。修正して下さい'
    pk = self.request.POST['run']
    run = Run.objects.get(order=pk)
    context ={
      'form':form,
      'msg':msg,
      'pk':pk,
      'runs':run,
    }
    return render(self.request,'score/create.html' ,context)

#確定のボタンを押したときの処理
class Determine(CreateView):
  form_class = CountForm
  model = Count  
  def calc_score(self):
    m1_score = int(self.request.POST.get("m1_count")) * 100
    m2_score = int(self.request.POST.get("m2_count")) * 400
    m2h_score = int(self.request.POST.get("m2h_count")) * 200
    m3_score = int(self.request.POST.get("m3_count")) * 400
    m3h_score = int(self.request.POST.get("m3h_count")) * 200
    m4_score = int(self.request.POST.get("m4_count")) * 600
    m4h_score = int(self.request.POST.get("m4h_count")) * 300
    m6_score = int(self.request.POST.get("m6_count")) * 1000
    m6h_score = int(self.request.POST.get("m6h_count")) * 500
    bonus1_score = int(self.request.POST.get("bonus1_count")) * 200
    if(self.request.POST.get("m5_count") == True):
      m5_score = 500
    else:
      m5_score = 0
      
    if(self.request.POST.get("m7_count") == True):
      m7_score = 500
    else:
      m7_score = 0
      
    if(self.request.POST.get("bonus2_count") == True):
      bonus2_score = 200
    else:
      bonus2_score = 0
      
    total = m1_score + m2_score + m2h_score +m3_score + m3h_score + m4_score + m4h_score + m6_score + m6h_score + m5_score +m7_score + bonus2_score + bonus1_score
    
    score = {
      'm1_score':m1_score,
      'm2_score':m2_score,
      'm2h_score':m2h_score,
      'm3_score':m3_score,
      'm3h_score':m3h_score,
      'm4_score':m4_score,
      'm4h_score':m4h_score,
      'm5_score':m5_score,
      'm6_score':m6_score,
      'm6h_score':m6h_score,
      'm7_score':m7_score,
      'bonus1_score':bonus1_score,
      'bonus2_score':bonus2_score,
      'total':total,
    }
    return score
  #フォームが通らなかった時の処理
  def form_invalid(self, form):
    context ={'form':form}
    return render(self.request, 'score/create.html', context)
  
  #確定ボタンを押したときに保存する処理
  def post(self,request,*args, **kwargs):
    #Base64データから画像に変換
    data_url = self.request.POST.get("img")
    splitString = data_url.split(',')
    img_binary = base64.b64decode(splitString[1])
    filename=self.request.POST.get("team")+'_'+self.request.POST.get("run_n")+'.png'
    #Countの保存
    pk = self.request.POST.get('run')
    run = Run.objects.get(order=pk)
    count_obj = Count(
      run=run,
      m1_count=self.request.POST.get("m1_count"),
      m2_count=self.request.POST.get("m2_count"),
      m2h_count=self.request.POST.get("m2h_count"),
      m3_count=self.request.POST.get("m3_count"),
      m3h_count=self.request.POST.get("m3h_count"),
      m4_count=self.request.POST.get("m4_count"),
      m4h_count=self.request.POST.get("m4h_count"),
      m5_count=self.request.POST.get("m5_count"),
      m6_count=self.request.POST.get("m6_count"),
      m6h_count=self.request.POST.get("m6h_count"),
      m7_count=self.request.POST.get("m7_count"),
      bonus1_count=self.request.POST.get("bonus1_count"),
      bonus2_count=self.request.POST.get("bonus2_count"),
      perfect=self.request.POST.get("perfect"),
      clear_time=self.request.POST.get("clear_time"), 
      sign=ContentFile(img_binary, filename)
    )
    count_obj.save()
    #Scoreの保存
    score = self.calc_score()
    score_obj = Score(
      run=run,
      m1_score=score["m1_score"],
      m2_score=score["m2_score"],
      m2h_score=score["m2h_score"],
      m3_score=score["m3_score"],
      m3h_score=score["m3h_score"],
      m4_score=score["m4_score"],
      m4h_score=score["m4h_score"],
      m5_score=score["m5_score"],
      m6_score=score["m6_score"],
      m6h_score=score["m6h_score"],
      m7_score=score["m7_score"],
      bonus1_score=score["bonus1_score"],
      bonus2_score=score["bonus2_score"],
      total=score["total"]
      )
    score_obj.save()
    #RunのStatusの変更
    run_obj = Run.objects.get(order=pk)
    run_obj.status = "走行後"
    run_obj.save()
    return render(request, 'score/rank.html')

#第1走行、第2走行かかわらずに順番に並べた順位を表示する
class Rank(TemplateView):
  Model = Run
  template_name = 'score/rank.html'     
  def get_context_data(self, **kwargs):
    score = Score.objects.order_by("run__rank")
    context = super().get_context_data(**kwargs)
    context.update({
      'scores':score,
    })
    return context
  
  def get(self, request, *args, **kwargs):
    scores = Score.objects.all().order_by('total','-run__count__clear_time').reverse()
    for i,score in enumerate(scores,1):
      order = score.run.order
      team = Run.objects.get(order=order)
      team.rank = i      
      team.save()
     
    return super().get(request, *args, **kwargs)

#審判の台本を表示する
class Libretto(TemplateView):
  template_name ='score/libretto.html'

#データを消すRunテーブルのデータは消さない
#特にページがあるわけではない関数ベースで作ってもいいかも
class ResetData(TemplateView):
  template_name = 'score/index.html'
  model = Run
  def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        run = Run.objects.all().order_by("order")
        context['runs'] = run
        #context['form'] = CountForm
        return context
      
  def get(self, request, *args, **kwargs):
    count =Count.objects.all()
    count.delete()
    score = Score.objects.all()
    score.delete()
    ranks = Run.objects.all()
    for rank in ranks:
      rank.rank = None
      rank.save()
    runs = Run.objects.all()
    for run in runs:
      run.status = "走行前"
      run.save()
    return super().get(request, *args, **kwargs)

#総合的な順位を表示する
#チームの得点の高いほうを採用して、再度順位付けをしている。順位はFor文のカウンターで出しているのでモデルには保存されていない
#手元にほしい場合はデータの書き出しを行うとExcelファイルが発行される
class Rank2(TemplateView):
  Model = Run
  template_name = 'score/rank2.html'
  def get_context_data(self, **kwargs):
    rank_list = self.remove_double()
    context = super().get_context_data(**kwargs)
    context.update({
      'rank_list':rank_list,
    })
    return context
  
  def remove_double(self, **kwargs):
    teams = Run.objects.all()
    d_TEAMS =[]
    TEAMS =[]
    for i in teams:
      d_TEAMS.append(i.team)
    
    TEAMS = set(d_TEAMS)
    rank_list =[]
    scores = Score.objects.select_related("run").order_by("-total","run__count__clear_time")
    for team_name in TEAMS:
      team = scores.filter(run__team=team_name)
      if team.count() == 0:
        continue
      elif team.count() == 1:
        rank_list.append([team[0],team[0].total,team[0].run.count.clear_time])
      elif team.count() == 2:
        if team[0].total >= team[1].total:
          rank_list.append([team[0],team[0].total,team[0].run.count.clear_time])
        else:
          rank_list.append([team[1],team[1].total,team[1].run.count.clear_time])
    sorted_data = sorted(rank_list,key=itemgetter(1), reverse=True)
    sorted(sorted_data,key=itemgetter(2))
    return sorted_data

#第1走行、第2走行をまとめたデータを出力する
#各ミッションのカウントや、得点を確認する場合はこちらを使ってExcelデータをはかせる
def ExportData(request):
  wb = openpyxl.Workbook()
  response = HttpResponse(content_type='application/vnd.ms-excel')
  response['Content-Disposition'] = 'attachment; filename=SRC_result.xlsx'
  
  results = Count.objects.all()
  form = CountForm()
  i=1
  for count in results:
    i=2
    sheet_title = str(count.run.team + count.run.run)
    ws = wb.create_sheet(title=sheet_title)
    for field in form:
      ws.cell(i, 1).value = field.label
      i+=1
    ws.cell(18, 1).value = "走行回数"
    ws.cell(19, 1).value ="合計得点"
    ws.cell(20, 1).value ="全体での順位"
    
    ws.cell(1,2).value ="回数"
    ws.cell(1,3).value ="点数"
    #カウントの入力
    ws.cell(2, 2).value = count.run.team
    ws.cell(3, 2).value = count.m1_count
    ws.cell(4, 2).value = count.m2_count
    ws.cell(5, 2).value = count.m2h_count
    ws.cell(6, 2).value = count.m3_count
    ws.cell(7, 2).value = count.m3h_count
    ws.cell(8, 2).value = count.m4_count
    ws.cell(9, 2).value = count.m4h_count
    ws.cell(10, 2).value = count.m5_count
    ws.cell(11, 2).value = count.m6_count
    ws.cell(12, 2).value = count.m6h_count
    ws.cell(13, 2).value = count.m7_count
    ws.cell(14, 2).value = count.bonus1_count
    ws.cell(15, 2).value = count.bonus2_count
    ws.cell(16, 2).value = count.perfect
    ws.cell(17, 2).value = count.clear_time
    ws.cell(18, 2).value = count.run.run
    ws.cell(20, 2).value =str(count.run.rank) +'位'
    #得点の入力
    ws.cell(3, 3).value = count.run.score.m1_score
    ws.cell(4, 3).value = count.run.score.m2_score
    ws.cell(5, 3).value = count.run.score.m2h_score
    ws.cell(6, 3).value = count.run.score.m3_score
    ws.cell(7, 3).value = count.run.score.m3h_score
    ws.cell(8, 3).value = count.run.score.m4_score
    ws.cell(9, 3).value = count.run.score.m4h_score
    ws.cell(10, 3).value = count.run.score.m5_score
    ws.cell(11, 3).value = count.run.score.m6_score
    ws.cell(12, 3).value = count.run.score.m6h_score
    ws.cell(13, 3).value = count.run.score.m7_score
    ws.cell(14, 3).value = count.run.score.bonus1_score
    ws.cell(15, 3).value = count.run.score.bonus2_score
    ws.cell(19, 3).value = count.run.score.total
    
    #セルの幅調整
    ws.column_dimensions['A'].width =35
    ws.column_dimensions['B'].width =25
    ws.column_dimensions['C'].width =25

    
  wb.remove(wb["Sheet"])
  wb.save(response)
  return response

#総合的なデータを出力する
#最終的な結果はこちらを見たほうが早い、各ミッションのカウントや、得点は載せていない
def ExportData2(request):
  wb = openpyxl.Workbook()
  response = HttpResponse(content_type='application/vnd.ms-excel')
  response['Content-Disposition'] = 'attachment; filename=SRC_result_all.xlsx'
  rank2 = Rank2()
  rank_list = Rank2.remove_double(rank2)
  ws = wb.create_sheet(title="総合順位")
  ws.cell(1, 1).value = "順位"
  ws.cell(1, 2).value = "回数"
  ws.cell(1, 3).value = "チーム名"
  ws.cell(1, 4).value = "得点"
  ws.cell(1, 5).value = "クリアタイム"
  for i,result in enumerate(rank_list,1):
      ws.cell(i+1, 1).value = i
      ws.cell(i+1, 2).value = result[0].run.run
      ws.cell(i+1, 3).value = result[0].run.team
      ws.cell(i+1, 4).value = result[0].total
      ws.cell(i+1, 5).value = result[0].run.count.clear_time
  ws.column_dimensions['C'].width =25
  ws.column_dimensions['E'].width =25
  wb.remove(wb["Sheet"])
  wb.save(response)
  return response

#保護者用のページ、走行順が確認できる
#スタート時間も書いておくべきか？その場合はテーブルに時間を追加する必要がある
#得点をいじれないようにリンクは消している、上のメニューバーも消す
class Order(TemplateView):
  template_name = 'score/order.html'
  model = Run
  def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        run = Run.objects.all().order_by("order")
        context['runs'] = run
        return context

  def get(self, request, *args, **kwargs):
    return super().get(request, *args, **kwargs)
  
